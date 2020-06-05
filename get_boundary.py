# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression
from curve_fit import polynomial_fit

df_object = pd.read_csv('data/nds-sync-object-14.csv', encoding="ISO-8859-1")
df_vehicle = pd.read_csv('data/nds-sync-vehicle-14.csv', encoding="ISO-8859-1")

df_object_16 = pd.read_csv('data/nds-sync-object-16.csv', encoding="ISO-8859-1")
df_vehicle_16 = pd.read_csv('data/nds-sync-vehicle-16.csv', encoding="ISO-8859-1")

workbook = load_workbook(u'data/ScenariosLabeling2tianda.xlsx')
booksheet = workbook.active

AdditionalDescription = booksheet['H']
OPosition = booksheet['L']
CBehavior = booksheet['G']

speed = []
a = []
obj_speed = []
obj_a = []

lc = []
rc = []

thw = []

for x in range(207):
    # if (CBehavior[x].value == "变道向左" and OPosition[x].value=="左后") or (CBehavior[x].value == "变道向右" and OPosition[x].value=="右后"):
    if (AdditionalDescription[x].value == "变道超车" and OPosition[x].value == "前"):
        start_time = booksheet.cell(row=x + 1, column=9).value
        endTime = booksheet.cell(row=x + 1, column=10).value
        id = booksheet.cell(row=x + 1, column=16).value

        speed.extend(df_vehicle[(df_vehicle.Time == start_time + 1)]['Vehicle Speed[KPH]'].values)
        a.extend(df_vehicle[(df_vehicle.Time == start_time + 1)]['Acceleration-x[m/s2].1'].values)

        lc.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['LC'].values)
        rc.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['RC'].values)

        obj_speed.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['VXAbs'].values)
        obj_a.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['AXAbs'].values)

        thw.extend(df_object[(df_object.Time == start_time + 1) & (df_object.PublicID == id)]['THW'].values)

for x in range(912, 1703, 1):
    # if (CBehavior[x].value == "变道向左" and OPosition[x].value=="左后") or (CBehavior[x].value == "变道向右" and OPosition[x].value=="右后"):
    if (AdditionalDescription[x].value == "变道超车" and OPosition[x].value == "前"):
        start_time = booksheet.cell(row=x + 1, column=9).value
        endTime = booksheet.cell(row=x + 1, column=10).value
        id = booksheet.cell(row=x + 1, column=16).value

        speed.extend(df_vehicle_16[(df_vehicle_16.Time == start_time + 1)]['Vehicle Speed[KPH]'].values)
        a.extend(df_vehicle_16[(df_vehicle_16.Time == start_time + 1)]['Acceleration-x[m/s2].1'].values)

        lc.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['LC'].values)
        rc.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['RC'].values)

        obj_speed.extend(
            df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['VXAbs'].values)
        obj_a.extend(
            df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['AXAbs'].values)

        thw.extend(df_object_16[(df_object_16.Time == start_time + 1) & (df_object_16.PublicID == id)]['THW'].values)
    if len(speed) != len(obj_speed):
        speed.pop()

speed = np.array(speed)
a = np.array(a)
obj_speed = np.array(obj_speed)
obj_a = np.array(obj_a)
lc = np.array(lc)
rc = np.array(rc)
thw = np.array(thw)

a = a.astype(np.float64)
speed = speed.astype(np.float64)
obj_speed = obj_speed.astype(np.float64)
obj_a = obj_a.astype(np.float64)
# lc = lc.astype(np.float64)
# rc = rc.astype(np.float64)

# print(thw)
# thw = thw.astype(np.float64)

del_pos = []
for n in range(speed.shape[0]):
    if thw[n] == 'na':
        del_pos.append(n)

print(del_pos)

speed = np.delete(speed, del_pos, axis=0)
a = np.delete(a, del_pos, axis=0)
obj_speed = np.delete(obj_speed, del_pos, axis=0)
obj_a = np.delete(obj_a, del_pos, axis=0)
lc = np.delete(lc, del_pos, axis=0)
rc = np.delete(rc, del_pos, axis=0)
thw = np.delete(thw, del_pos, axis=0)

# lc = lc.astype(np.float64)
# rc = rc.astype(np.float64)

thw = thw.astype(np.float64)

distance = thw * speed

print(speed)
print(np.max(speed))
print(np.min(speed))

speed1 = [65, 75, 85, 95, 105]
# speed1 = [62.5, 67.5, 72.5, 77.5, 82.5, 87.5, 92.5, 97.5, 102.5, 107.5, 112.5, 117.5]

a_max = []
a_min = []
obj_speed_max = []
obj_speed_min = []
obj_a_max = []
obj_a_min = []
distance_max = []
distance_min = []
lc_max = []
lc_min = []
rc_max = []
rc_min = []

start = 60

for i in range(5):
    print(i)
    # if a[np.where((speed>start+i*10) & (speed<start+i*10+10))].shape[0] == 0:
    #     a_max.append(0)
    #     a_min.append(0)
    # else:
    a_max.append(a[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].max())
    a_min.append(a[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].min())

    obj_speed_max.append(obj_speed[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].max())
    obj_speed_min.append(obj_speed[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].min())

    obj_a_max.append(obj_a[np.where((obj_speed > start + i * 10) & (obj_speed < start + i * 10 + 10))].max())
    obj_a_min.append(obj_a[np.where((obj_speed > start + i * 10) & (obj_speed < start + i * 10 + 10))].min())

    # lc_max.append(lc[np.where((speed>start+i*10) & (speed<start+i*10+10))].max())
    # lc_min.append(lc[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].min())
    #
    # rc_max.append(rc[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].max())
    # rc_min.append(rc[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].min())

    distance_max.append(distance[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].max())
    distance_min.append(distance[np.where((speed > start + i * 10) & (speed < start + i * 10 + 10))].min())

# for i in range(5):
#     if distance[np.where((speed-obj_speed > start + i * 5) & (speed-obj_speed < start + i * 5 + 5))].shape[0] == 0:
#         print(i)
#         continue
#     else:
#         distance_max.append(distance[np.where((speed-obj_speed > start + i * 5) & (speed-obj_speed < start + i * 5 + 5))].max())
#         distance_min.append(distance[np.where((speed-obj_speed > start + i * 5) & (speed-obj_speed < start + i * 5 + 5))].min())


speed1 = np.array(speed1).reshape(-1, 1)
a_max = np.array(a_max).reshape(-1, 1)
a_min = np.array(a_min).reshape(-1, 1)
obj_speed_max = np.array(obj_speed_max).reshape(-1, 1)
obj_speed_min = np.array(obj_speed_min).reshape(-1, 1)
obj_a_max = np.array(obj_a_max).reshape(-1, 1)
obj_a_min = np.array(obj_a_min).reshape(-1, 1)
distance_max = np.array(distance_max).reshape(-1, 1)
distance_min = np.array(distance_min).reshape(-1, 1)
# lc_max = np.array(a_max).reshape(-1,1)
# lc_min = np.array(a_min).reshape(-1,1)
#
# rc_max = np.array(a_max).reshape(-1,1)
# rc_min = np.array(a_min).reshape(-1,1)


# 设置颜色和多项式的阶数
degrees = [1, 2, 3, 4]
colors = ['b', 'y', 'g', 'r']
font = {'family': 'Times New Roman', 'weight': 'normal', 'size': 15, }

# 设置画布大小
plt.figure(figsize=(20, 10))

# 速度-加速度
plt.subplot(121)
plt.scatter(speed, a, color='black')
plt.ylabel('a', font)
plt.xlabel('speed', font)

plt.subplot(122)
plt.xlim([55, 125])
polynomial_fit(speed1, a_max, degrees, colors)
polynomial_fit(speed1, a_min, degrees, colors, 0)

plt.scatter(speed1, a_max, color='black', marker='x')
plt.scatter(speed1, a_min, color='black')
plt.scatter(speed, a, color='grey')
plt.ylabel('a', font)
plt.xlabel('speed', font)

plt.show()

plt.figure(figsize=(20, 10))
# 速度-目标车速度
plt.subplot(121)
plt.scatter(speed, obj_speed, color='black')
plt.ylabel('obj_speed', font)
plt.xlabel('speed', font)

plt.subplot(122)
plt.xlim([55, 125])

polynomial_fit(speed1, obj_speed_max, degrees, colors)
polynomial_fit(speed1, obj_speed_min, degrees, colors, 0)

plt.scatter(speed1, obj_speed_max, color='black', marker='x')
plt.scatter(speed1, obj_speed_min, color='black')
plt.scatter(speed, obj_speed, color='gray')
plt.ylabel('obj_speed', font)
plt.xlabel('speed', font)

plt.show()

# 设置画布大小
plt.figure(figsize=(20, 10))
# 速度-相对距离
plt.subplot(121)
plt.scatter(speed, distance, color='black')
plt.ylabel('distance', font)
plt.xlabel('speed', font)

plt.subplot(122)
plt.xlim([55, 125])
polynomial_fit(speed1, distance_max, degrees, colors)
polynomial_fit(speed1, distance_min, degrees, colors, 0)
#
#
plt.scatter(speed1, distance_max, color='black', marker='x')
plt.scatter(speed1, distance_min, color='black')

plt.scatter(speed, distance, color='grey')

# plt.scatter(speed, distance,  color='black')
plt.ylabel('distance', font)
plt.xlabel('speed', font)

plt.show()
plt.figure(figsize=(20, 10))

# 目标车速度-目标车加速度
plt.subplot(121)
plt.scatter(speed, obj_a, color='black')
plt.ylabel('obj_a', font)
plt.xlabel('obj_speed', font)

plt.subplot(122)
plt.xlim([55, 125])
polynomial_fit(speed1, obj_a_max, degrees, colors)
polynomial_fit(speed1, obj_a_min, degrees, colors, 0)

plt.scatter(speed1, obj_a_max, color='black', marker='x')
plt.scatter(speed1, obj_a_min, color='black')
plt.scatter(speed, obj_a, color='grey')

# plt.scatter(speed, obj_a,  color='black')
plt.ylabel('obj_a', font)
plt.xlabel('obj_speed', font)

plt.show()
